"""Strictly lagged features from the National Bank of Kazakhstan archive."""
from __future__ import annotations

import datetime as dt
import hashlib
from pathlib import Path

import numpy as np
from openpyxl import load_workbook

from ml.data import Series


DATA = Path("data/external_kazakhstan_nbk_rub_usd_cny_2016_2026.xlsx")
LAGS = (1, 2, 5, 10, 20)


def load_kazakh_nbk(path: Path = DATA) -> tuple[dict[str, Series], str]:
    payload = path.read_bytes()
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() for value in next(rows)]
    columns = {code: headers.index(code) for code in ("RUB", "USD", "CNY")}
    records = {code: {} for code in columns}
    for values in rows:
        if not values[0]:
            continue
        day = dt.datetime.strptime(str(values[0]), "%d.%m.%Y").date()
        for code, column in columns.items():
            nominal = float(values[column - 1])
            value = float(values[column])
            if nominal > 0 and value > 0:
                records[code][day] = value / nominal
    workbook.close()
    result = {}
    for code, mapping in records.items():
        ordered = sorted(mapping.items())
        if not ordered:
            raise ValueError(f"Kazakh NBK archive has no {code} observations")
        result[code] = Series(
            code,
            np.asarray([row[0] for row in ordered], dtype=object),
            np.asarray([row[1] for row in ordered], dtype=float),
        )
    return result, hashlib.sha256(payload).hexdigest()


def _last(series: Series, day: dt.date, strict: bool) -> tuple[int, float, int]:
    stop = int(np.searchsorted(series.dates, day, side="left" if strict else "right"))
    if not stop:
        return stop, np.nan, 999
    return stop, float(series.values[stop - 1]), (day - series.dates[stop - 1]).days


def _basis(a: float, b: float) -> float:
    return float(np.log(a / b) * 10000.0) if a > 0 and b > 0 else 0.0


def _ret(values: np.ndarray, stop: int, lag: int) -> float:
    if stop <= lag:
        return 0.0
    return float(np.log(values[stop - 1] / values[stop - 1 - lag]) * 10000.0)


def build_kazakh_nbk_features(index, target_series, cbr_reference, local):
    rows = []
    names = None
    for _currency, _position, day in index:
        rstop, rub, rage = _last(local["RUB"], day, strict=True)
        ustop, usd, uage = _last(local["USD"], day, strict=True)
        cstop, cny, cage = _last(local["CNY"], day, strict=True)
        _kstop, cbr_kzt, kage = _last(target_series["KZT"], day, strict=False)
        _ustop, cbr_usd, _ = _last(cbr_reference["USD"], day, strict=False)
        _cstop, cbr_cny, _ = _last(cbr_reference["CNY"], day, strict=False)
        available = all(np.isfinite(value) and value > 0 for value in (
            rub, usd, cny, cbr_kzt, cbr_usd, cbr_cny,
        ))
        direct = 1.0 / rub if available else 0.0
        via_usd = cbr_usd / usd if available else 0.0
        via_cny = cbr_cny / cny if available else 0.0
        implied = (direct, via_usd, via_cny)
        values = []
        row_names = []
        for label, value in zip(("direct", "usd", "cny"), implied):
            values.extend((value, _basis(value, cbr_kzt)))
            row_names.extend((
                f"kazakh_nbk_{label}_implied_kzt_rub",
                f"kazakh_nbk_{label}_basis_bps",
            ))
        consensus = float(np.exp(np.mean(np.log(implied)))) if available else 0.0
        values.extend((
            consensus, _basis(consensus, cbr_kzt),
            _basis(direct, via_usd), _basis(direct, via_cny),
            _basis(via_usd, via_cny),
        ))
        row_names.extend((
            "kazakh_nbk_consensus_implied_kzt_rub",
            "kazakh_nbk_consensus_basis_bps",
            "kazakh_nbk_direct_minus_usd_bps",
            "kazakh_nbk_direct_minus_cny_bps",
            "kazakh_nbk_usd_minus_cny_bps",
        ))
        for code, stop in (("rub", rstop), ("usd", ustop), ("cny", cstop)):
            for lag in LAGS:
                values.append(_ret(local[code.upper()].values, stop, lag))
                row_names.append(f"kazakh_nbk_{code}_quote_ret_{lag}")
        values.extend((
            float(min(rage, 30)), float(min(uage, 30)),
            float(min(cage, 30)), float(min(kage, 30)), float(not available),
        ))
        row_names.extend((
            "kazakh_nbk_rub_age_days", "kazakh_nbk_usd_age_days",
            "kazakh_nbk_cny_age_days", "kazakh_nbk_cbr_kzt_age_days",
            "kazakh_nbk_missing",
        ))
        rows.append(values)
        if names is None:
            names = row_names
        elif names != row_names:
            raise AssertionError("Kazakh NBK feature schema changed")
    matrix = np.asarray(rows, dtype=np.float32)
    if not np.all(np.isfinite(matrix)):
        raise ValueError("non-finite Kazakh NBK feature")
    return matrix, names or []


def causality_check(
    index, target_series, cbr_reference, local,
    cutoff=dt.date(2025, 6, 30),
):
    full, names = build_kazakh_nbk_features(
        index, target_series, cbr_reference, local,
    )
    changed = {}
    for code, series in local.items():
        values = series.values.copy()
        future = series.dates >= cutoff
        values[future] *= np.linspace(2.0, 50.0, int(future.sum()))
        changed[code] = Series(code, series.dates.copy(), values)
    altered, altered_names = build_kazakh_nbk_features(
        index, target_series, cbr_reference, changed,
    )
    past = np.asarray([row[2] <= cutoff for row in index])
    if names != altered_names or not np.array_equal(full[past], altered[past]):
        raise AssertionError("future Kazakh NBK value changed a past feature")
    if not np.any(full[~past] != altered[~past]):
        raise AssertionError("future Kazakh NBK corruption did not affect future rows")
    return True
